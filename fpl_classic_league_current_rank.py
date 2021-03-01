#Import Libraries
import requests
import openpyxl

#Create Excel Worksheet
global g_w
wb = openpyxl.Workbook()
sheet1 = wb.create_sheet(index=1, title='Classic League Ranks')

#Load data from API
url9 = 'https://fantasy.premierleague.com/api/entry/806959/'
json_info = requests.get(url9).json()


#League Rank header
sheet1.merge_cells('A1:B1')
sheet1['A1'].value = 'League Rank History'

clrow = 2
num_of_leagues = len(json_info['leagues']['classic'])
clheader = ['League Name', 'Rank']
for leaguecolumn in range(2):
    sheet1.cell(row=clrow, column=leaguecolumn + 1).value = str(clheader[leaguecolumn])
for each in json_info['leagues']['classic']:
    leaguename = each['name']
    leagueposition = each['entry_rank']
    league_data = [leaguename, leagueposition]
    clrow = clrow + 1
    for clcol in range(2):
        sheet1.cell(row=clrow, column=clcol + 1).value = league_data[clcol]


#Save Workbook
wb.save("ClassicFPL.xlsx")

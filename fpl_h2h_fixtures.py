#Import Libraries
import requests
import openpyxl

#Function to generate url endpoint for api
def GetEndpoints():
    apis = []
    baseEndpoint = "https://fantasy.premierleague.com/api/leagues-h2h-matches/league/556449/"    
    for i in range(1, 8):
        if i == 1 :
          print(baseEndpoint)
          apis.append(baseEndpoint)
        else:
            apiEndpoint = baseEndpoint + "?page=" + str(i)
            print(apiEndpoint)
            apis.append(apiEndpoint)
    return apis

#Function to get responses from API
def GetAPIResponse(APIEndpointURL):
    response = requests.get(APIEndpointURL).json()
    return response
    
#Convert API response to JSON
def ConvertToJson(data):
    response = data.json()

#Get all response jsons and write to excel
def GetAndWriteDataToExcel():
    apis = GetEndpoints()
    for api in apis:
        response = GetAPIResponse(api)
        writeJsonToExcel(sheet1,response)

#Function to remove repetition
def writeJsonToExcel(sheet1,json_history):
  global rownum
  for each in json_history["results"]:
    ide = each['id']
    Gameweek = each['event']
    Home = each['entry_1_name']
    Away = each['entry_2_name']
    HomePoints = each['entry_1_points']
    AwayPoints = each['entry_2_points']
    HomeTotal = each['entry_1_total']
    AwayTotal = each['entry_2_total']

    Rowvalue = sheet1.cell(row=2, column=3).value
    if not Rowvalue :
        rownum = 2

    sheet1.cell(row=rownum, column=3).value = ide
    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Home
    sheet1.cell(row=rownum, column=6).value = HomePoints
    sheet1.cell(row=rownum, column=7).value = HomeTotal
    sheet1.cell(row=rownum, column=8).value = Away
    sheet1.cell(row=rownum, column=9).value = AwayPoints
    sheet1.cell(row=rownum, column=10).value = AwayTotal

    rownum +=1

#Running all of the functionality from here
#Create excel workbook,create sheets
global g_w,wb,sheet0,sheet1  
wb = openpyxl.Workbook()
sheet0 = wb.create_sheet(index=0, title='Intro Page')
sheet1 = wb.create_sheet(index=1, title='Fantasy')
sheet2 = wb.create_sheet(index=2, title='Other')


#Create Read me sheet
sheet0['B2'].value = 'Welcome to FPL data'

header1 = ['id', 'GW', 'HomeTeamName', 'GWPoints','H2HPts', 'AwayTeamName', 'GWPoints', 'H2HPts']
headerrow = 1
for key in range(8):
    sheet1.cell(row=headerrow, column=key + 3).value = str(header1[key])
# Import gameweek history and insert data in sheet

#Get all apis, get data from each and write to excel
GetAndWriteDataToExcel()
#Save Workbook
wb.save("FPL2022Fixtures.xlsx")




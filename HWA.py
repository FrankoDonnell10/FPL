#Import Libraries
import requests
import pandas as pd
import openpyxl

#Create Excel Worksheet
global g_w
wb = openpyxl.Workbook()
sheet0 = wb.create_sheet(index=0, title='Read_Me')
sheet1 = wb.create_sheet(index=1, title='2019_2020')

#Create Read me sheet
sheet0['B2'].value = 'Welcome to FPL data'

#Import API data and retrieve team name
url = "https://fantasy.premierleague.com/api/leagues-h2h-matches/league/585139/"
url2 = "https://fantasy.premierleague.com/api/leagues-h2h-matches/league/585139/?page=2"
url3 = "https://fantasy.premierleague.com/api/leagues-h2h-matches/league/585139/?page=3"
url4 = "https://fantasy.premierleague.com/api/leagues-h2h-matches/league/585139/?page=4"
url5 = "https://fantasy.premierleague.com/api/leagues-h2h-matches/league/585139/?page=5"
url6 = "https://fantasy.premierleague.com/api/leagues-h2h-matches/league/585139/?page=6"
url7 = "https://fantasy.premierleague.com/api/leagues-h2h-matches/league/585139/?page=7"
json_history = requests.get(url).json()
json_history2 = requests.get(url2).json()
json_history3 = requests.get(url3).json()
json_history4 = requests.get(url4).json()
json_history5 = requests.get(url5).json()
json_history6 = requests.get(url6).json()
json_history7 = requests.get(url7).json()
   
# Import gameweek history and insert data in sheet
header1 = ['id', 'GW', 'TeamName', 'GWPoints','H2HPts',]
headerrow = 1
for key in range(5):
    sheet1.cell(row=headerrow, column=key + 3).value = str(header1[key])

# To make a list of overall rank for inserting change in rank symbols
#declaring rownum as 2, incrementing it by 1 at the end of each loop
rownum = 2
for each in json_history["results"]:
    ide = each['id']
    Gameweek = each['event']
    Home = each['entry_1_name']
    Away = each['entry_2_name']
    HomePoints = each['entry_1_points']
    AwayPoints = each['entry_2_points']
    HomeTotal = each['entry_1_total']
    AwayTotal = each['entry_2_total']


    sheet1.cell(row=rownum, column=3).value = ide
    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Home
    sheet1.cell(row=rownum, column=6).value = HomePoints
    sheet1.cell(row=rownum, column=7).value = HomeTotal
    rownum +=1

    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Away
    sheet1.cell(row=rownum, column=6).value = AwayPoints
    sheet1.cell(row=rownum, column=7).value = AwayTotal

    rownum +=1

#rownum = 52
for each in json_history2["results"]:
    ide = each['id']
    Gameweek = each['event']
    Home = each['entry_1_name']
    Away = each['entry_2_name']
    HomePoints = each['entry_1_points']
    AwayPoints = each['entry_2_points']
    HomeTotal = each['entry_1_total']
    AwayTotal = each['entry_2_total']


    sheet1.cell(row=rownum, column=3).value = ide
    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Home
    sheet1.cell(row=rownum, column=6).value = HomePoints
    sheet1.cell(row=rownum, column=7).value = HomeTotal
    rownum +=1

    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Away
    sheet1.cell(row=rownum, column=6).value = AwayPoints
    sheet1.cell(row=rownum, column=7).value = AwayTotal

    rownum +=1

#rownum = 102
for each in json_history3["results"]:
    ide = each['id']
    Gameweek = each['event']
    Home = each['entry_1_name']
    Away = each['entry_2_name']
    HomePoints = each['entry_1_points']
    AwayPoints = each['entry_2_points']
    HomeTotal = each['entry_1_total']
    AwayTotal = each['entry_2_total']


    sheet1.cell(row=rownum, column=3).value = ide
    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Home
    sheet1.cell(row=rownum, column=6).value = HomePoints
    sheet1.cell(row=rownum, column=7).value = HomeTotal
    rownum +=1

    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Away
    sheet1.cell(row=rownum, column=6).value = AwayPoints
    sheet1.cell(row=rownum, column=7).value = AwayTotal

    rownum +=1
    
#rownum = 152
for each in json_history4["results"]:
    ide = each['id']
    Gameweek = each['event']
    Home = each['entry_1_name']
    Away = each['entry_2_name']
    HomePoints = each['entry_1_points']
    AwayPoints = each['entry_2_points']
    HomeTotal = each['entry_1_total']
    AwayTotal = each['entry_2_total']


    sheet1.cell(row=rownum, column=3).value = ide
    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Home
    sheet1.cell(row=rownum, column=6).value = HomePoints
    sheet1.cell(row=rownum, column=7).value = HomeTotal
    rownum +=1

    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Away
    sheet1.cell(row=rownum, column=6).value = AwayPoints
    sheet1.cell(row=rownum, column=7).value = AwayTotal

    rownum +=1

#rownum = 202
for each in json_history5["results"]:
    ide = each['id']
    Gameweek = each['event']
    Home = each['entry_1_name']
    Away = each['entry_2_name']
    HomePoints = each['entry_1_points']
    AwayPoints = each['entry_2_points']
    HomeTotal = each['entry_1_total']
    AwayTotal = each['entry_2_total']


    sheet1.cell(row=rownum, column=3).value = ide
    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Home
    sheet1.cell(row=rownum, column=6).value = HomePoints
    sheet1.cell(row=rownum, column=7).value = HomeTotal
    rownum +=1

    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Away
    sheet1.cell(row=rownum, column=6).value = AwayPoints
    sheet1.cell(row=rownum, column=7).value = AwayTotal

    rownum +=1

#rownum = 252
for each in json_history6["results"]:
    ide = each['id']
    Gameweek = each['event']
    Home = each['entry_1_name']
    Away = each['entry_2_name']
    HomePoints = each['entry_1_points']
    AwayPoints = each['entry_2_points']
    HomeTotal = each['entry_1_total']
    AwayTotal = each['entry_2_total']


    sheet1.cell(row=rownum, column=3).value = ide
    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Home
    sheet1.cell(row=rownum, column=6).value = HomePoints
    sheet1.cell(row=rownum, column=7).value = HomeTotal
    rownum +=1

    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Away
    sheet1.cell(row=rownum, column=6).value = AwayPoints
    sheet1.cell(row=rownum, column=7).value = AwayTotal

    rownum +=1

#rownum = 302
for each in json_history7["results"]:
    ide = each['id']
    Gameweek = each['event']
    Home = each['entry_1_name']
    Away = each['entry_2_name']
    HomePoints = each['entry_1_points']
    AwayPoints = each['entry_2_points']
    HomeTotal = each['entry_1_total']
    AwayTotal = each['entry_2_total']


    sheet1.cell(row=rownum, column=3).value = ide
    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Home
    sheet1.cell(row=rownum, column=6).value = HomePoints
    sheet1.cell(row=rownum, column=7).value = HomeTotal
    rownum +=1

    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Away
    sheet1.cell(row=rownum, column=6).value = AwayPoints
    sheet1.cell(row=rownum, column=7).value = AwayTotal

    rownum +=1

#Save Workbook
wb.save("Hereweare.xlsx")

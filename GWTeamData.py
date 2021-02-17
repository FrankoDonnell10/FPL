#Import Libraries
import requests
import pandas as pd
import openpyxl

#Create Excel Worksheet
global g_w
wb = openpyxl.Workbook()
sheet0 = wb.create_sheet(index=0, title='Read_Me')
sheet1 = wb.create_sheet(index=1, title='2019_2020')

#Create Read me sheet
sheet0['B2'].value = 'Hey all. This excel file is the result of a python script'

#Import API data and retrieve team name       
url1 = "https://fantasy.premierleague.com/api/entry/806959/history/"
url2 = "https://fantasy.premierleague.com/api/bootstrap-static/"
json_history = requests.get(url1).json()
json_live = requests.get(url2).json()
num_of_gw = len(json_history['current'])
participants = json_live['total_players']

## Import league data and team name from new url due to change in FPL api since 2019-2020 Season
url3 = 'https://fantasy.premierleague.com/api/entry/806959/'
json_info = requests.get(url3).json()
team_name = json_info['name']

# Import gameweek history and insert data in sheet
header1 = ['GW', 'GP', 'GW AVG', 'GW HS', 'PB', 'TM', 'TC', 'GR', 'PGR', 'OP', 'OR', 'POR', 'Position', 'TV']
headerrow = 1
for key in range(14):
    sheet1.cell(row=headerrow, column=key + 3).value = str(header1[key])

o_r1 = []  # To make a list of overall rank for inserting change in rank symbols
for each in json_history["current"]:
    g_w = each['event']
    points = each['points']
    p_b = each['points_on_bench']
    t_m = each['event_transfers']
    t_c = each['event_transfers_cost']
    g_w_r = each['rank']
    o_r = each['overall_rank']
    t_v = each['value']
    o_r1.append(o_r)    # This is for creating rank symbols in the excel sheet
    o_p = each['total_points']

    p_g_r = 100 - (((participants - g_w_r) / participants) * 100)
    p_o_r = 100 - (((participants - o_r) / participants) * 100)
    p_o_s = 1  # placeholder which will be replaced later down the code

    history_list = [g_w, points, p_b, t_m, t_c, g_w_r, p_g_r, o_p, o_r, p_o_r, p_o_s, t_v / 10]
    for rownum in range(g_w + 1, g_w + 2):
        sheet1.cell(row=rownum, column=3).value = g_w
    for rownum in range(g_w + 1, g_w + 2):
        sheet1.cell(row=rownum, column=4).value = points
    for rownum in range(g_w + 1, g_w + 2):
        for key in range(2, 12):
            sheet1.cell(row=rownum, column=key + 5).value = history_list[key]

#Save Workbook
wb.save("Fpl.Team.Info.xlsx")
wb.save("Fpl.Team.Info.csv")





